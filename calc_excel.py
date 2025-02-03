#import Jeddah_steel_interface
import datetime
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

JB = Workbook()
JB = load_workbook('JeddahSteel.xlsx')
JS = JB.active
def write_data(e1, e2, e3, width, utilize, wastage, weightMM, Tpipe, DCno, order_no, comment, company, coil):
    x = 2
    while JS.cell(row=x, column=1).value != None:
        print(JS.cell(row=x,column=1).value)
        x += 1
        if x > 1048576:
            messagebox.showerror("Error", "Excel row limit exceeded!")
            return

    cost = (int(utilize) - int(wastage)) * int(width) * 10

    T = datetime.datetime.now()
    T = str(T)
    T_array = T.split(" ")

    JS.cell(row=x, column=1).value = T_array[0]
    JS.cell(row=x, column=2).value = T_array[1][0:8]
    JS.cell(row=x, column=3).value = e1.get()
    if e2.get() == "":
        JS.cell(row=x, column=4).value = "-"
    if e3.get() == "":
        JS.cell(row=x, column=5).value = "-"
    JS.cell(row=x, column=6).value = company
    JS.cell(row=x, column=7).value = int(order_no)
    JS.cell(row=x, column=8).value = int(DCno)
    JS.cell(row=x, column=9).value = int(width)
    JS.cell(row=x, column=10).value = int(weightMM)
    JS.cell(row=x, column=11).value = int(Tpipe)
    JS.cell(row=x, column=12).value = coil
    JS.cell(row=x, column=13).value = int(cost)
    JS.cell(row=x, column=14).value = comment
    JB.save('JeddahSteel.xlsx')


def search_date(search_item):
    x = 2
    # Header with fixed-width columns
    item_string = f"{'Time':<10}{'Job Number':<12}{'Supplier Name':<15}{'Coil Type':<12}{'Cost':<10}\n"

    while JS.cell(row=x, column=1).value is not None:
        if JS.cell(row=x, column=1).value == search_item:
            item_string += (
                f"{str(JS.cell(row=x, column=2).value):<10}"  # Time
                f"{str(JS.cell(row=x, column=3).value):<12}"  # Job Number
                f"{str(JS.cell(row=x, column=6).value):<15}"  # Supplier Name
                f"{str(JS.cell(row=x, column=12).value):<12}"  # Coil Type
                f"{str(JS.cell(row=x, column=13).value):<10}"  # Cost
                "\n"
            )
        x += 1  # Increment the row index to prevent infinite loop

    return item_string

